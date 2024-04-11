"Detect  New Pools Created on Solana Raydium DEX"

#MAnually see transactions of new pairs GThUX1Atko4tqhN2NaiTazWSeFWMuiUvfFnyJyUghFMJ under spl transfer section
from save_tokens_to_db import  remove_token_from_database, update_database_with_new_token

from time import sleep
import logging
import pandas as pd
from datetime import datetime
from datetime import timedelta

import os
from openpyxl import load_workbook
import json
import asyncio
from typing import List, AsyncIterator, Tuple, Iterator
from asyncstdlib import enumerate
import time
import httpx
from typing import Any
from httpx import AsyncClient
from solders.pubkey import Pubkey
from solders.rpc.config import RpcTransactionLogsFilterMentions

from solana.rpc.websocket_api import connect
from solana.rpc.commitment import Finalized
from solana.rpc.api import Client
from solana.exceptions import SolanaRpcException
from websockets.exceptions import ConnectionClosedError, ProtocolError
from httpx import AsyncClient
# Type hinting imports
from solana.rpc.commitment import Commitment
from solana.rpc.websocket_api import SolanaWsClientProtocol
from solders.rpc.responses import RpcLogsResponse, SubscriptionResult, LogsNotification, GetTransactionResp
from solders.signature import Signature
from solders.transaction_status import UiPartiallyDecodedInstruction, ParsedInstruction
from dotenv import load_dotenv
load_dotenv()

# Raydium Liquidity Pool V4
RaydiumLPV4 = "675kPX9MHTjS2zt1qfr1NYHuzeLXfQM9H24wFSUt1Mp8"
URI = "https://mainnet.helius-rpc.com/?api-key=b0030426-49da-4a2a-ab94-ef8961452c7c"  # "https://api.devnet.solana.com" | "https://api.mainnet-beta.solana.com"
WSS = "wss://mainnet.helius-rpc.com/?api-key=b0030426-49da-4a2a-ab94-ef8961452c7c"  # "wss://api.devnet.solana.com" | "wss://api.mainnet-beta.solana.com"
solana_client = Client(URI)
# Raydium function call name, look at raydium-amm/program/src/instruction.rs
log_instruction = "initialize2"
seen_signatures = set()
seen_tokens_filename = 'seen_tokens.txt'


# Init logging
logging.basicConfig(filename='app.log', filemode='a', level=logging.DEBUG)
# Writes responses from socket to messages.json
# Writes responses from http req to  transactions.json

filename = os.getenv('FILE_PATH')
all_tokens_filename = 'all_tokens.txt'


def save_seen_tokens(tokens, filename):
    """Save the seen tokens list to a file."""
    with open(filename, 'w') as file:
        for token in tokens:
            file.write(f"{token}\n")

def load_all_tokens(filename):
    """Load all tokens from a file."""
    try:
        with open(filename, 'r') as file:
            tokens = file.read().splitlines()
            return tokens
    except FileNotFoundError:
        return []
def load_seen_tokens(filename):
    """Load seen tokens from a file."""
    try:
        with open(filename, 'r') as file:
            tokens = file.read().splitlines()
            return tokens
    except FileNotFoundError:
        return []
seen_tokens = load_seen_tokens(seen_tokens_filename)
all_tokens = load_all_tokens(all_tokens_filename)



high_volume_tokens = {}
high_volume_file =  os.getenv('VOLUME_FILE_PATH')

def load_high_volume_tokens(filename="high_volume_tokens.json"):
    """Load the high volume tokens dictionary from a file."""
    try:
        with open(filename, 'r') as f:
            tokens = json.load(f)
            for token, details in tokens.items():
                if 'timestamp' in details:
                    details['timestamp'] = datetime.fromisoformat(details['timestamp'])
            return tokens
    except FileNotFoundError:
        print("No previous high volume tokens file found. Starting fresh.")
        return {}
    except Exception as e:
        print(f"Failed to load high volume tokens: {e}")
        return {}   

def remove_token_from_high_volume(token_address):
    if token_address in high_volume_tokens:
        del high_volume_tokens[token_address]
        save_high_volume_tokens(high_volume_tokens)
        print(f"Token {token_address} removed from high volume tracking due to price drop.")
def save_high_volume_tokens(tokens, filename="high_volume_tokens.json"):
    """Save the high volume tokens dictionary to a file."""
    try:
        with open(filename, 'w') as f:
            json.dump(tokens, f, default=str)
    except Exception as e:
        print(f"Failed to save high volume tokens: {e}")
async def periodically_save_state():
    while True:
        await asyncio.sleep(60)  
        save_high_volume_tokens(high_volume_tokens)   

async def websocket_listener_task():
    async for websocket in connect(WSS):
        try:
            subscription_id = await subscribe_to_logs(
                websocket,
                RpcTransactionLogsFilterMentions(RaydiumLPV4),
                Finalized
            )
            async for i, signature in enumerate(process_messages(websocket, log_instruction)):  # type: ignore
                logging.info(f"{i=}")
                try:
                    get_tokens(signature, RaydiumLPV4)
                except Exception as e:
                    logging.exception(e)
                    sleep(5)
                    continue
        except Exception as e:
            logging.exception(e)
            continue
        except KeyboardInterrupt:
            if websocket:
                await websocket.logs_unsubscribe(subscription_id)
                break
            
async def main():
    try:
        listener_task = asyncio.create_task(robust_websocket_listener_task())
        dexscreener_task = asyncio.create_task(call_dexscreener_api())
        high_volume_task = asyncio.create_task(DexScreenerHighVolume())

        save_state_task = asyncio.create_task(periodically_save_state())
        await asyncio.gather(listener_task, dexscreener_task,high_volume_task, save_state_task)
    except KeyboardInterrupt:
        print("Shutdown requested...cleaning up")
        await cleanup()
        print("Cleanup complete. Exiting now.")
        return
    
async def cleanup():
    """Perform cleanup operations before script exit."""
    # Save high volume tokens to file
    try:
        save_high_volume_tokens(high_volume_tokens)
        print("High volume tokens saved successfully.")
    except Exception as e:
        print(f"Failed to save high volume tokens: {e}")

    # Save seen tokens to file
    try:
        save_seen_tokens(seen_tokens, seen_tokens_filename)
        print("Seen tokens saved successfully.")
    except Exception as e:
        print(f"Failed to save seen tokens: {e}")
async def subscribe_to_logs(websocket: SolanaWsClientProtocol, 
                            mentions: RpcTransactionLogsFilterMentions,
                            commitment: Commitment) -> int:
    await websocket.logs_subscribe(
        filter_=mentions,
        commitment=commitment
    )
    first_resp = await websocket.recv()
    return get_subscription_id(first_resp)  # type: ignore


def get_subscription_id(response: SubscriptionResult) -> int:
    return response[0].result


async def process_messages(websocket: SolanaWsClientProtocol,
                           instruction: str) -> AsyncIterator[Signature]:
    """Async generator, main websocket's loop"""
    async for idx, msg in enumerate(websocket):
        value = get_msg_value(msg)
        if not idx % 100:
            pass
            # print(f"{idx=}")
        for log in value.logs:
            if instruction not in log:
                continue
            # Start logging
            logging.info(value.signature)
            logging.info(log)
            # Logging to messages.json
            with open("messages.json", 'a', encoding='utf-8') as raw_messages:  
                raw_messages.write(f"signature: {value.signature} \n")
                raw_messages.write(msg[0].to_json())
                raw_messages.write("\n ########## \n")
            # End logging
            yield value.signature


def get_msg_value(msg: List[LogsNotification]) -> RpcLogsResponse:
    return msg[0].result.value

def get_tokens(signature: Signature, RaydiumLPV4: Pubkey) -> None:
    """httpx.HTTPStatusError: Client error '429 Too Many Requests' 
    for url 'https://api.mainnet-beta.solana.com'
    For more information check: https://httpstatuses.com/429

    """
    global seen_signatures, seen_tokens
    if signature in seen_signatures:
        # If we have already seen this signature, skip processing.
        logging.info(f"Duplicate transaction skipped: {signature}")
        return
    else:
        # Mark this signature as seen.
        seen_signatures.add(signature)
    try:
        # Attempt to get the transaction
        transaction = solana_client.get_transaction(
            signature,
            encoding='jsonParsed',
            max_supported_transaction_version=0
        )
        
        # Check if transaction is not None
        if transaction is None:
            logging.error(f'No transaction found for signature: {signature}')
            return
        
        # Start logging to transactions.json
        with open('transactions.json', 'a', encoding='utf-8') as raw_transactions:
            raw_transactions.write(f'signature: {signature}\\n')
            raw_transactions.write(transaction.to_json())        
            raw_transactions.write('\\n ########## \\n')
        # End logging
        
        instructions = get_instructions(transaction)
        filtered_instructions = instructions_with_program_id(instructions, RaydiumLPV4)
        logging.info(filtered_instructions)
        for instruction in filtered_instructions:
            tokens = get_tokens_info(instruction)
            print_table(tokens)
            # print(f'True, https://solscan.io/tx/{signature}')

    except AttributeError as e:
        # Catching attribute errors if the 'transaction' is None or doesn't have the expected attributes
        logging.exception(f'AttributeError with signature {signature}: {e}')
    except httpx.HTTPStatusError as e:
        # Handling HTTPStatusError specifically for '429 Too Many Requests'
        if e.response.status_code == 429:
            logging.warning('429 Too Many Requests: The request is being rate limited.')
            # Implementing a basic exponential backoff strategy
            wait = 5  # Starting with a 1-minute backoff
            max_attempts = 5
            for i in range(max_attempts):
                time.sleep(wait)
                try:
                    # Retry the transaction fetch
                    transaction = solana_client.get_transaction(
                        signature,
                        encoding='jsonParsed',
                        max_supported_transaction_version=0
                    )
                    # If success, break out of the retry loop
                    if transaction is not None:
                        break
                except httpx.HTTPStatusError as e:
                    # If still getting '429', increase the wait time
                    wait *= 2
                    logging.warning(f'Retrying after {wait} seconds...')
                except Exception as e:
                    # Handle any other exceptions that occur during the retry
                    logging.exception(f'An unexpected error occurred: {e}')
                    break
            else:
                logging.error('Max retries reached, the transaction could not be fetched.')
    except Exception as e:
        # Catching all other exceptions
        logging.exception(f'An unexpected error occurred with signature {signature}: {e}')

def get_instructions(
    transaction: GetTransactionResp
) -> List[UiPartiallyDecodedInstruction | ParsedInstruction]:
    instructions = transaction \
                   .value \
                   .transaction \
                   .transaction \
                   .message \
                   .instructions
    return instructions


def instructions_with_program_id(
    instructions: List[UiPartiallyDecodedInstruction | ParsedInstruction],
    program_id: str
) -> Iterator[UiPartiallyDecodedInstruction | ParsedInstruction]:
    return (instruction for instruction in instructions
            if instruction.program_id == program_id)
    
def save_all_tokens(tokens, filename):
    """Save the all tokens list to a file."""
    with open(filename, 'w') as file:
        for token in tokens:
            file.write(f"{token}\n")
def get_tokens_info(instruction):
    SOL_TOKEN_ADDRESS = "So11111111111111111111111111111111111111112"
    global seen_tokens
    global all_tokens

    accounts = instruction.accounts
    Pair = accounts[4]
    Token0 = accounts[8]
    Token1 = accounts[9]

    if str(Token0) == SOL_TOKEN_ADDRESS:
        token_to_add = str(Token1)
    elif str(Token1) == SOL_TOKEN_ADDRESS:
        token_to_add = str(Token0)
    else:
        token_to_add = str(Token0)

    if token_to_add not in seen_tokens and token_to_add != SOL_TOKEN_ADDRESS:
        seen_tokens.append(token_to_add)
        update_database_with_new_token(token_to_add)
        logging.info(f"Token {token_to_add} added to seen_tokens.")
    
    if token_to_add not in all_tokens:
        all_tokens.append(token_to_add)
        save_all_tokens(all_tokens, all_tokens_filename)

    if token_to_add not in high_volume_tokens:
        high_volume_tokens[token_to_add] = {'timestamp': datetime.now(), 'met_criteria': False}

    return (Token0, Token1, Pair)



async def DexScreenerHighVolume():
    while True:
        current_time = datetime.now()
        print(f"Starting API call at {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        tokens_to_evaluate = []
        tokens_to_remove = []
        tokens_for_excel = []

        for token, details in high_volume_tokens.items():
            time_since_added = current_time - details['timestamp']
            if time_since_added > timedelta(minutes=15):
                if details['met_criteria']:
                    tokens_for_excel.append(token)  
                if not details['met_criteria']:
                    print(f"Removing token {token} - did not meet criteria within 15 minutes")
                    tokens_to_remove.append(token)
                continue
            tokens_to_evaluate.append(token)

        for token in tokens_to_remove:
            remove_token_from_high_volume(token)

        if tokens_for_excel:
            token_addresses = ','.join(tokens_for_excel)
            url = f"https://api.dexscreener.com/latest/dex/tokens/{token_addresses}"
            try:
                async with AsyncClient() as client:
                    response = await client.get(url)
                    data = response.json()

                    if 'pairs' in data:
                        append_to_high_volume_excel(data['pairs'], high_volume_file)
            except Exception as e:
                logging.error(f"Error fetching token data for Excel appending: {e}")

        if tokens_to_evaluate:
            token_addresses = ','.join(tokens_to_evaluate)
            url = f"https://api.dexscreener.com/latest/dex/tokens/{token_addresses}"
            try:
                async with AsyncClient() as client:
                    response = await client.get(url)
                    data = response.json()
                    print(f"API response data: {data}")

                    for pair in data.get('pairs', []):
                        token_address = pair['baseToken'].get('address')
                        if token_address not in tokens_to_evaluate:
                            continue 

                        volume_m5 = float(pair['volume']['m5'])
                        price_change_h1 = float(pair['priceChange']['h1'])
                        liquidity_usd = float(pair['liquidity']['usd'])

                        if volume_m5 > 150000 and liquidity_usd > 10000 and pair['volume']['h1']<pair.get('fdv'):
                            if price_change_h1 < -40:
                                print(f"Token {token_address} has met criteria but now removed due to price drop of {price_change_h1}% in the last hour.")
                                remove_token_from_high_volume(token_address)
                            else:
                                high_volume_tokens[token_address]['met_criteria'] = True
                                print(f"Token {token_address} has met all criteria without a significant price drop in the last hour.")


            except Exception as e:
                logging.error(f"Error fetching data from DexScreener for high volume tokens: {e}")

        save_high_volume_tokens(high_volume_tokens)

        await asyncio.sleep(300)
        
def append_to_high_volume_excel(pairs, high_volume_file):
    pairs_data = []
    print(f'{pairs}pairs')
    for pair in pairs:
        token_address = pair['baseToken'].get('address')
        if token_address in high_volume_tokens and high_volume_tokens[token_address]['met_criteria']:
            formatted_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            pair_created_at = datetime.fromtimestamp(pair.get('pairCreatedAt', 0) / 1000).strftime('%Y-%m-%d %H:%M:%S')

            pair_info = {
                    'Timestamp': formatted_timestamp,
                    'DEX ID': pair.get('dexId'),
                    'Base Token Address': pair['baseToken'].get('address'),
                    'Base Token Symbol': pair['baseToken'].get('symbol'),
                    'Price USD': pair.get('priceUsd'),
                    '5m Buys (Txn)': pair.get('txns', {}).get('m5', {}).get('buys'),
                    '5m Sells (Txn)': pair.get('txns', {}).get('m5', {}).get('sells'),
                    '1h Buys (Txn)': pair.get('txns', {}).get('h1', {}).get('buys'),
                    '1h Sells (Txn)': pair.get('txns', {}).get('h1', {}).get('sells'),
                    '6h Buys (Txn)': pair.get('txns', {}).get('h6', {}).get('buys'),
                    '6h Sells (Txn)': pair.get('txns', {}).get('h6', {}).get('sells'),
                    '24h Buys (Txn)': pair.get('txns', {}).get('h24', {}).get('buys'),
                    '24h Sells (Txn)': pair.get('txns', {}).get('h24', {}).get('sells'),
                    '24h Buy-Sells (Txn)': pair.get('txns', {}).get('h24', {}).get('buys', 0) - pair.get('txns', {}).get('h24', {}).get('sells', 0),
                    'Volume 5m': pair.get('volume', {}).get('m5'),
                    'Volume 1h': pair.get('volume', {}).get('h1'),
                    'Volume 6h': pair.get('volume', {}).get('h6'),
                    'Price Change 5m': pair.get('priceChange', {}).get('m5'),
                    'Price Change 1h': pair.get('priceChange', {}).get('h1'),
                    'Price Change 6h': pair.get('priceChange', {}).get('h6'),
                    'FDV': pair.get('fdv'),
                    'Price Change 24h': pair.get('priceChange', {}).get('h24'),
                    'Volume 24h': pair.get('volume', {}).get('h24'),
                    'Liquidity USD': pair.get('liquidity', {}).get('usd'),
                    'pair_created_at' : pair_created_at,
                    'Website': pair.get('website_url')
            }
            pairs_data.append(pair_info)
    
    df = pd.DataFrame(pairs_data)
    if not df.empty:
        try:
            if os.path.exists(high_volume_file) and os.path.getsize(high_volume_file) > 0:
                with pd.ExcelWriter(high_volume_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)
            else:
                df.to_excel(high_volume_file, index=False)
                print('created excel file')
            print("Saved to high volume token info success!")
        except Exception as e:
            print(e)

def print_table(tokens: Tuple[Pubkey, Pubkey, Pubkey]) -> None:
    data = [
        {'Token_Index': 'Token0', 'Account Public Key': tokens[0]},  # Token0
        {'Token_Index': 'Token1', 'Account Public Key': tokens[1]},  # Token1
        {'Token_Index': 'LP Pair', 'Account Public Key': tokens[2]}  # LP Pair
    ]
    print("============NEW POOL DETECTED====================")
    header = ["Token_Index", "Account Public Key"]
    print("│".join(f" {col.ljust(15)} " for col in header))
    print("|".rjust(18))
    for row in data:
        print("│".join(f" {str(row[col]).ljust(15)} " for col in header))
async def call_dexscreener_api():
    batch_size = 30
    while True:
        for i in range(0, len(seen_tokens), batch_size):
            current_batch = seen_tokens[i:i + batch_size]
            if current_batch:
                token_addresses = ','.join(str(token) for token in current_batch)
                url = f"https://api.dexscreener.com/latest/dex/tokens/{token_addresses}"
 
                try:
                    async with AsyncClient() as client:
                        response = await client.get(url)
                        response.raise_for_status()
                        data = response.json()
                        for pair in data.get('pairs', []):
                            token_address = pair['baseToken'].get('address')
                            volume_5m = pair.get('volume', {}).get('m5', 0)
                            price_change_5m = pair.get('priceChange', {}).get('m5', 0) 
                            volume_1h = pair.get('volume', {}).get('h1', 0)
                            price_change_h24 = pair.get('priceChange', {}).get('h24', 0)
                            if  price_change_h24 < -90:
                                if token_address in seen_tokens:
                                    seen_tokens.remove(token_address)
                                    # print(f"Removed token {token_address} due to price change")

                        append_to_excel(data, filename) 

                except Exception as e:
                    logging.error(f"Error fetching data from DexScreener: {e}")
                await asyncio.sleep(1)  
        
        await asyncio.sleep(60)

def append_to_excel(data, filename):

    pairs_data = []
    for pair in data.get('pairs', []):
        formatted_timestamp = datetime.now().strftime('%m/%d/%y %H:%M')
        pair_created_at = datetime.fromtimestamp(pair.get('pairCreatedAt', 0) / 1000).strftime('%m/%d/%y %H:%M')

        pair_info = {
            'Timestamp': formatted_timestamp,
            'DEX ID': pair.get('dexId'),
            'Base Token Address': pair['baseToken'].get('address'),
            'Base Token Symbol': pair['baseToken'].get('symbol'),
            'Price USD': pair.get('priceUsd'),
            '5m Buys (Txn)': pair.get('txns', {}).get('m5', {}).get('buys'),
            '5m Sells (Txn)': pair.get('txns', {}).get('m5', {}).get('sells'),
            '1h Buys (Txn)': pair.get('txns', {}).get('h1', {}).get('buys'),
            '1h Sells (Txn)': pair.get('txns', {}).get('h1', {}).get('sells'),
            '6h Buys (Txn)': pair.get('txns', {}).get('h6', {}).get('buys'),
            '6h Sells (Txn)': pair.get('txns', {}).get('h6', {}).get('sells'),
            '24h Buys (Txn)': pair.get('txns', {}).get('h24', {}).get('buys'),
            '24h Sells (Txn)': pair.get('txns', {}).get('h24', {}).get('sells'),
            '24h Buy-Sells (Txn)': pair.get('txns', {}).get('h24', {}).get('buys', 0) - pair.get('txns', {}).get('h24', {}).get('sells', 0),
            'Volume 5m': pair.get('volume', {}).get('m5'),
            'Volume 1h': pair.get('volume', {}).get('h1'),
            'Volume 6h': pair.get('volume', {}).get('h6'),
            'Price Change 5m': pair.get('priceChange', {}).get('m5'),
            'Price Change 1h': pair.get('priceChange', {}).get('h1'),
            'Price Change 6h': pair.get('priceChange', {}).get('h6'),
            'FDV': pair.get('fdv'),
            'Price Change 24h': pair.get('priceChange', {}).get('h24'),
            'Volume 24h': pair.get('volume', {}).get('h24'),
            'Liquidity USD': pair.get('liquidity', {}).get('usd'),
            'pair_created_at' : pair_created_at,
            'Website': pair.get('website_url')
        }
        pairs_data.append(pair_info)
    
    df = pd.DataFrame(pairs_data)
    try:
        if os.path.exists(filename):
            with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        else:
            df.to_excel(filename, index=False)
    except Exception as e:
        print(e)






async def robust_websocket_listener_task(attempts=3, delay=10):
    attempt = 0
    while attempt < attempts:
        try:
            await websocket_listener_task()
            break  
        except Exception as e:
            logging.exception(f"websocket_listener_task failed: {e}")
            attempt += 1
            logging.info(f"Retrying websocket_listener_task, attempt {attempt}")
            await asyncio.sleep(delay)  #
    else:
        logging.error("websocket_listener_task failed after maximum retry attempts.")


if __name__ == "__main__":
    RaydiumLPV4 = Pubkey.from_string(RaydiumLPV4)
    high_volume_tokens = load_high_volume_tokens()
    try:
        asyncio.run(main())
    except Exception as e:
        print(f"An error occurred: {e}")


# def append_to_high_volume_excel(data, high_volume_file):
#     pairs_data = []
#     print(f'calling append{data},{high_volume_file},')
#     for pair in data.get('pairs', []):
#         token_address = pair['baseToken'].get('address')
#         if token_address in high_volume_tokens and high_volume_tokens[token_address]['met_criteria']:
#             formatted_timestamp = datetime.now().strftime('%m/%d/%y %H:%M')
#             pair_created_at = datetime.fromtimestamp(pair.get('pairCreatedAt', 0) / 1000).strftime('%m/%d/%y %H:%M')

#             pair_info = {
#                     'Timestamp': formatted_timestamp,
#                     'DEX ID': pair.get('dexId'),
#                     'Base Token Address': pair['baseToken'].get('address'),
#                     'Base Token Symbol': pair['baseToken'].get('symbol'),
#                     'Price USD': pair.get('priceUsd'),
#                     '5m Buys (Txn)': pair.get('txns', {}).get('m5', {}).get('buys'),
#                     '5m Sells (Txn)': pair.get('txns', {}).get('m5', {}).get('sells'),
#                     '1h Buys (Txn)': pair.get('txns', {}).get('h1', {}).get('buys'),
#                     '1h Sells (Txn)': pair.get('txns', {}).get('h1', {}).get('sells'),
#                     '6h Buys (Txn)': pair.get('txns', {}).get('h6', {}).get('buys'),
#                     '6h Sells (Txn)': pair.get('txns', {}).get('h6', {}).get('sells'),
#                     '24h Buys (Txn)': pair.get('txns', {}).get('h24', {}).get('buys'),
#                     '24h Sells (Txn)': pair.get('txns', {}).get('h24', {}).get('sells'),
#                     '24h Buy-Sells (Txn)': pair.get('txns', {}).get('h24', {}).get('buys', 0) - pair.get('txns', {}).get('h24', {}).get('sells', 0),
#                     'Volume 5m': pair.get('volume', {}).get('m5'),
#                     'Volume 1h': pair.get('volume', {}).get('h1'),
#                     'Volume 6h': pair.get('volume', {}).get('h6'),
#                     'Price Change 5m': pair.get('priceChange', {}).get('m5'),
#                     'Price Change 1h': pair.get('priceChange', {}).get('h1'),
#                     'Price Change 6h': pair.get('priceChange', {}).get('h6'),
#                     'FDV': pair.get('fdv'),
#                     'Price Change 24h': pair.get('priceChange', {}).get('h24'),
#                     'Volume 24h': pair.get('volume', {}).get('h24'),
#                     'Liquidity USD': pair.get('liquidity', {}).get('usd'),
#                     'pair_created_at' : pair_created_at,
#                     'Website': pair.get('website_url')
#             }
#             pairs_data.append(pair_info)
#     df = pd.DataFrame(pairs_data)
#     try:
#         if os.path.exists(high_volume_file) and os.path.getsize(high_volume_file) > 0:
#             with pd.ExcelWriter(high_volume_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#                 try:
#                     last_row = pd.read_excel(high_volume_file).shape[0]
#                 except ValueError:
#                     last_row = 0

#                 df.to_excel(writer, sheet_name='Sheet1', startrow=last_row + 1, index=False, header=False)
#                 print("Saved to high volume token info success!")

#         else:
#             df.to_excel(high_volume_file, index=False)
#             print("Saved to high volume token info success!")
    
#     except Exception as e:
#         print(e)

# async def DexScreenerHighVolume():
#     while True:
#         current_time = datetime.datetime.now()
#         tokens_to_process = [token for token, details in high_volume_tokens.items() if details['met_criteria']]
#         if tokens_to_process:
#             token_addresses = ','.join(tokens_to_process)
#             url = f"https://api.dexscreener.com/latest/dex/tokens/{token_addresses}"
#             print(url)
#             try:
#                 async with AsyncClient() as client:
#                     response = await client.get(url)
#                     response.raise_for_status()
#                     data = response.json()
#                     # print(data)
#                     pairs_meeting_criteria = [
#                         pair for pair in data.get('pairs', [])
#                         if pair['baseToken'].get('address') in tokens_to_process
#                         and float(pair.get('volume', {}).get('m5', 0)) > 150
#                     ]
#                     if pairs_meeting_criteria:
#                         append_to_high_volume_excel(pairs_meeting_criteria, high_volume_file)
#                     try:
#                         for pair in data.get('pairs', []):
#                             token_address = pair['baseToken'].get('address')
#                             if token_address in tokens_to_process:
#                                 print(token_address)
#                                 pairs_meeting_criteria = [pair for pair in data.get('pairs', []) if pair['baseToken'].get('address') in tokens_to_process]
#                         append_to_high_volume_excel(pairs_meeting_criteria, high_volume_file)
#                     except Exception as e:
#                         logging.error(f"Error savign high volume: {e}")
#             except Exception as e:
#                 logging.error(f"Error fetching data from DexScreener for high volume tokens: {e}")
#         await asyncio.sleep(5)
